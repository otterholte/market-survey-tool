"""
Market Survey Analysis Tool (MSAT) Generator
Generates a professional Excel workbook for student housing prelease tracking.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo


def create_styles():
    """Create reusable styles for the workbook."""
    # Colors
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_row_fill = PatternFill(start_color="D6E3F8", end_color="D6E3F8", fill_type="solid")
    accent_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    light_accent_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    
    # Fonts
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")
    body_font = Font(name="Segoe UI", size=10)
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7')
    )
    
    return {
        'header_fill': header_fill,
        'alt_row_fill': alt_row_fill,
        'accent_fill': accent_fill,
        'light_accent_fill': light_accent_fill,
        'header_font': header_font,
        'title_font': title_font,
        'body_font': body_font,
        'thin_border': thin_border
    }


def setup_market_averages_sheet(ws, styles):
    """Set up the Market Averages settings sheet."""
    ws.title = "Market Averages"
    
    # Title
    ws['B2'] = "Market Average Prelease Percentages"
    ws['B2'].font = styles['title_font']
    ws.merge_cells('B2:C2')
    
    # Instructions
    ws['B4'] = "Enter the market average prelease % for each bedroom type below."
    ws['B4'].font = Font(name="Segoe UI", size=10, italic=True, color="666666")
    ws.merge_cells('B4:D4')
    
    # Headers
    headers = ["Bedroom Type", "Prelease %"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = styles['thin_border']
    
    # Bedroom types with default values
    bedroom_types = [
        ("Studio", 0.45),
        ("1 BR", 0.50),
        ("2 BR", 0.55),
        ("3 BR", 0.60),
        ("4 BR", 0.55),
        ("5 BR", 0.50)
    ]
    
    for i, (bed_type, default_pct) in enumerate(bedroom_types, start=7):
        # Bedroom type label
        cell = ws.cell(row=i, column=2, value=bed_type)
        cell.font = styles['body_font']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = styles['thin_border']
        if i % 2 == 1:
            cell.fill = styles['alt_row_fill']
        
        # Prelease percentage
        cell = ws.cell(row=i, column=3, value=default_pct)
        cell.font = styles['body_font']
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = styles['thin_border']
        if i % 2 == 1:
            cell.fill = styles['alt_row_fill']
    
    # Set column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    
    # Create named ranges for easy formula reference
    # These will be used in the Property Data sheet
    return {
        'Studio': '$C$7',
        '1 BR': '$C$8',
        '2 BR': '$C$9',
        '3 BR': '$C$10',
        '4 BR': '$C$11',
        '5 BR': '$C$12'
    }


def setup_property_data_sheet(ws, styles, market_avg_refs):
    """Set up the Property Data main input sheet."""
    ws.title = "Property Data"
    
    # Title
    ws['B2'] = "Property & Floorplan Data Entry"
    ws['B2'].font = styles['title_font']
    ws.merge_cells('B2:G2')
    
    # Instructions
    ws['B4'] = "Enter property data below. Leave 'Prelease %' blank to use market averages."
    ws['B4'].font = Font(name="Segoe UI", size=10, italic=True, color="666666")
    ws.merge_cells('B4:G4')
    
    # Headers
    headers = [
        ("Property Name", 25),
        ("Floorplan Name", 20),
        ("Bedrooms", 12),
        ("Units", 10),
        ("Prelease %", 12),
        ("Effective %", 12),
        ("Total Beds", 12),
        ("Leased Beds", 12)
    ]
    
    for col, (header, width) in enumerate(headers, start=2):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['thin_border']
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Set up data validation for Bedrooms column (dropdown)
    bedroom_options = '"Studio,1 BR,2 BR,3 BR,4 BR,5 BR"'
    dv = DataValidation(type="list", formula1=bedroom_options, allow_blank=True)
    dv.error = "Please select a valid bedroom type"
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Select bedroom type"
    dv.promptTitle = "Bedrooms"
    ws.add_data_validation(dv)
    
    # Add 100 rows for data entry with formulas
    for row in range(7, 107):
        # Property Name (B)
        cell = ws.cell(row=row, column=2, value="")
        cell.font = styles['body_font']
        cell.border = styles['thin_border']
        if row % 2 == 1:
            cell.fill = styles['alt_row_fill']
        
        # Floorplan Name (C)
        cell = ws.cell(row=row, column=3, value="")
        cell.font = styles['body_font']
        cell.border = styles['thin_border']
        if row % 2 == 1:
            cell.fill = styles['alt_row_fill']
        
        # Bedrooms (D) - with dropdown
        cell = ws.cell(row=row, column=4, value="")
        cell.font = styles['body_font']
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles['thin_border']
        if row % 2 == 1:
            cell.fill = styles['alt_row_fill']
        dv.add(cell)
        
        # Units (E)
        cell = ws.cell(row=row, column=5, value="")
        cell.font = styles['body_font']
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles['thin_border']
        if row % 2 == 1:
            cell.fill = styles['alt_row_fill']
        
        # Prelease % (F) - optional input
        cell = ws.cell(row=row, column=6, value="")
        cell.font = styles['body_font']
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles['thin_border']
        if row % 2 == 1:
            cell.fill = styles['alt_row_fill']
        
        # Effective % (G) - formula: use Prelease % if provided, else market average
        # VLOOKUP to get market average based on bedroom type
        formula = (
            f'=IF(F{row}<>"",F{row},'
            f'IF(D{row}="Studio",\'Market Averages\'!$C$7,'
            f'IF(D{row}="1 BR",\'Market Averages\'!$C$8,'
            f'IF(D{row}="2 BR",\'Market Averages\'!$C$9,'
            f'IF(D{row}="3 BR",\'Market Averages\'!$C$10,'
            f'IF(D{row}="4 BR",\'Market Averages\'!$C$11,'
            f'IF(D{row}="5 BR",\'Market Averages\'!$C$12,"")))))))'
        )
        cell = ws.cell(row=row, column=7, value=formula)
        cell.font = styles['body_font']
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles['thin_border']
        if row % 2 == 1:
            cell.fill = styles['alt_row_fill']
        
        # Total Beds (H) - formula: bedrooms * units
        beds_formula = (
            f'=IF(OR(D{row}="",E{row}=""),"",'
            f'IF(D{row}="Studio",1,'
            f'IF(D{row}="1 BR",1,'
            f'IF(D{row}="2 BR",2,'
            f'IF(D{row}="3 BR",3,'
            f'IF(D{row}="4 BR",4,'
            f'IF(D{row}="5 BR",5,0))))))*E{row})'
        )
        cell = ws.cell(row=row, column=8, value=beds_formula)
        cell.font = styles['body_font']
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles['thin_border']
        if row % 2 == 1:
            cell.fill = styles['alt_row_fill']
        
        # Leased Beds (I) - formula: total beds * effective %
        leased_formula = f'=IF(OR(H{row}="",G{row}=""),"",ROUND(H{row}*G{row},0))'
        cell = ws.cell(row=row, column=9, value=leased_formula)
        cell.font = Font(name="Segoe UI", size=10, bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles['thin_border']
        if row % 2 == 1:
            cell.fill = styles['alt_row_fill']
    
    # Add conditional formatting to highlight rows using market averages
    # (when Prelease % is blank but there's data)
    market_avg_rule = FormulaRule(
        formula=['AND($F7="",$D7<>"")'],
        fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    )
    ws.conditional_formatting.add('B7:I106', market_avg_rule)
    
    # Freeze panes (header row)
    ws.freeze_panes = 'B7'
    
    # Column A spacer
    ws.column_dimensions['A'].width = 3


def setup_report_sheet(ws, styles):
    """Set up the Leased Beds Report output sheet."""
    ws.title = "Leased Beds Report"
    
    # Title
    ws['B2'] = "Leased Beds Summary Report"
    ws['B2'].font = styles['title_font']
    ws.merge_cells('B2:F2')
    
    # Report date
    ws['B3'] = '=TODAY()'
    ws['B3'].font = Font(name="Segoe UI", size=10, italic=True, color="666666")
    ws['B3'].number_format = 'MMMM D, YYYY'
    
    # Instructions
    ws['B5'] = "This report auto-updates from Property Data. Copy this table for your reports."
    ws['B5'].font = Font(name="Segoe UI", size=10, italic=True, color="666666")
    ws.merge_cells('B5:F5')
    
    # Headers
    headers = [
        ("Property Name", 25),
        ("Floorplan", 18),
        ("Beds/Unit", 12),
        ("Total Beds", 12),
        ("Prelease %", 12),
        ("Leased Beds", 12)
    ]
    
    for col, (header, width) in enumerate(headers, start=2):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = styles['thin_border']
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add formulas that reference Property Data sheet
    for row in range(8, 108):
        data_row = row - 1  # Property Data starts at row 7
        
        # Property Name (B)
        cell = ws.cell(row=row, column=2, value=f'=IF(\'Property Data\'!B{data_row}="","",\'Property Data\'!B{data_row})')
        cell.font = styles['body_font']
        cell.border = styles['thin_border']
        if row % 2 == 0:
            cell.fill = styles['alt_row_fill']
        
        # Floorplan (C)
        cell = ws.cell(row=row, column=3, value=f'=IF(\'Property Data\'!C{data_row}="","",\'Property Data\'!C{data_row})')
        cell.font = styles['body_font']
        cell.border = styles['thin_border']
        if row % 2 == 0:
            cell.fill = styles['alt_row_fill']
        
        # Beds/Unit (D) - derived from bedroom type
        beds_formula = (
            f'=IF(\'Property Data\'!D{data_row}="","",'
            f'IF(\'Property Data\'!D{data_row}="Studio",1,'
            f'IF(\'Property Data\'!D{data_row}="1 BR",1,'
            f'IF(\'Property Data\'!D{data_row}="2 BR",2,'
            f'IF(\'Property Data\'!D{data_row}="3 BR",3,'
            f'IF(\'Property Data\'!D{data_row}="4 BR",4,'
            f'IF(\'Property Data\'!D{data_row}="5 BR",5,"")))))))'
        )
        cell = ws.cell(row=row, column=4, value=beds_formula)
        cell.font = styles['body_font']
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles['thin_border']
        if row % 2 == 0:
            cell.fill = styles['alt_row_fill']
        
        # Total Beds (E)
        cell = ws.cell(row=row, column=5, value=f'=IF(\'Property Data\'!H{data_row}="","",\'Property Data\'!H{data_row})')
        cell.font = styles['body_font']
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles['thin_border']
        if row % 2 == 0:
            cell.fill = styles['alt_row_fill']
        
        # Prelease % (F)
        cell = ws.cell(row=row, column=6, value=f'=IF(\'Property Data\'!G{data_row}="","",\'Property Data\'!G{data_row})')
        cell.font = styles['body_font']
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles['thin_border']
        if row % 2 == 0:
            cell.fill = styles['alt_row_fill']
        
        # Leased Beds (G)
        cell = ws.cell(row=row, column=7, value=f'=IF(\'Property Data\'!I{data_row}="","",\'Property Data\'!I{data_row})')
        cell.font = Font(name="Segoe UI", size=10, bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles['thin_border']
        if row % 2 == 0:
            cell.fill = styles['alt_row_fill']
    
    # Summary section
    summary_row = 110
    
    ws.cell(row=summary_row, column=2, value="TOTALS").font = Font(name="Segoe UI", size=11, bold=True, color="1F4E79")
    
    # Total Beds Sum
    cell = ws.cell(row=summary_row, column=5, value='=SUMIF(E8:E107,"<>""")')
    cell.font = Font(name="Segoe UI", size=11, bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = styles['light_accent_fill']
    cell.border = styles['thin_border']
    
    # Leased Beds Sum
    cell = ws.cell(row=summary_row, column=7, value='=SUMIF(G8:G107,"<>""")')
    cell.font = Font(name="Segoe UI", size=11, bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = styles['light_accent_fill']
    cell.border = styles['thin_border']
    
    # Overall Prelease %
    ws.cell(row=summary_row + 2, column=2, value="Overall Prelease:").font = Font(name="Segoe UI", size=11, bold=True, color="1F4E79")
    cell = ws.cell(row=summary_row + 2, column=3, value='=IF(E110=0,"",G110/E110)')
    cell.font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
    cell.number_format = '0.0%'
    
    # Freeze panes
    ws.freeze_panes = 'B8'
    
    # Column A spacer
    ws.column_dimensions['A'].width = 3


def add_sample_data(ws):
    """Add sample data to demonstrate the tool."""
    sample_data = [
        ("The Heights", "Studio Deluxe", "Studio", 24, 0.52),
        ("The Heights", "A1", "1 BR", 36, 0.52),
        ("The Heights", "B1", "2 BR", 48, 0.52),
        ("The Heights", "B2 Premium", "2 BR", 24, 0.52),
        ("The Heights", "C1", "3 BR", 32, 0.52),
        ("University Village", "Efficiency", "Studio", 20, ""),
        ("University Village", "One Bed", "1 BR", 40, ""),
        ("University Village", "Two Bed A", "2 BR", 60, ""),
        ("University Village", "Two Bed B", "2 BR", 30, ""),
        ("University Village", "Three Bed", "3 BR", 48, ""),
        ("University Village", "Four Bed", "4 BR", 24, ""),
        ("Campus Edge", "Studio", "Studio", 16, 0.48),
        ("Campus Edge", "1BR Classic", "1 BR", 32, 0.48),
        ("Campus Edge", "2BR Standard", "2 BR", 40, 0.48),
        ("Campus Edge", "3BR Townhome", "3 BR", 20, 0.48),
    ]
    
    for i, (prop, floorplan, beds, units, prelease) in enumerate(sample_data, start=7):
        ws.cell(row=i, column=2, value=prop)
        ws.cell(row=i, column=3, value=floorplan)
        ws.cell(row=i, column=4, value=beds)
        ws.cell(row=i, column=5, value=units)
        if prelease != "":
            ws.cell(row=i, column=6, value=prelease)


def generate_workbook(output_path="MarketSurvey.xlsx", include_sample_data=True):
    """Generate the complete Market Survey workbook."""
    print("Creating Market Survey Analysis Tool...")
    
    wb = Workbook()
    styles = create_styles()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # Create sheets
    ws_market = wb.create_sheet("Market Averages", 0)
    ws_data = wb.create_sheet("Property Data", 1)
    ws_report = wb.create_sheet("Leased Beds Report", 2)
    
    # Remove the default empty sheet
    wb.remove(default_sheet)
    
    # Set up each sheet
    print("  Setting up Market Averages sheet...")
    market_refs = setup_market_averages_sheet(ws_market, styles)
    
    print("  Setting up Property Data sheet...")
    setup_property_data_sheet(ws_data, styles, market_refs)
    
    print("  Setting up Leased Beds Report sheet...")
    setup_report_sheet(ws_report, styles)
    
    # Add sample data if requested
    if include_sample_data:
        print("  Adding sample data...")
        add_sample_data(ws_data)
    
    # Set Property Data as the active sheet
    wb.active = ws_data
    
    # Save the workbook
    print(f"  Saving to {output_path}...")
    wb.save(output_path)
    print(f"\nSuccess! Created {output_path}")
    print("\nHow to use:")
    print("  1. Open 'Market Averages' sheet to set your market default prelease %")
    print("  2. Enter property/floorplan data in 'Property Data' sheet")
    print("  3. View calculated results in 'Leased Beds Report' sheet")
    print("\nTip: Rows highlighted in yellow are using market averages (no property-specific % entered)")


if __name__ == "__main__":
    generate_workbook()

